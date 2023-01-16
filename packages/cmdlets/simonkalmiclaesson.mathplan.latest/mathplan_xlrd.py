try:
	from openpyxl import load_workbook,Workbook
except:
    os.system("python3 -m pip install openpyxl")
    from openpyxl import load_workbook

file = f"{CSScriptRoot}{os.sep}Matematik M2_2023.xlsx"

#Open excel workbook
workbook = load_workbook(filename=file)
new_workbook = Workbook()

#Select first sheet
worksheet = workbook.active
new_worksheet = workbook.active

#Initialize empty list for data
data = []

# Initialize a variable to keep track of the current week
week_num = None

# Iterate through hyperlinks
hyperlinks = {}
c = 0
for _ in worksheet.iter_rows():
    hyperlinks[c] = ""
for i in range(0,len(worksheet)):
    row = worksheet.iter_rows()[i]
    if row[7] != None:
        if row[7].hyperlink:
            string_name = str(row[7].value)
            hyperlinks[string_name] = row[7].hyperlink.target
        else:
            hyperlinks[str(i)] = ""


for i,row in enumerate(worksheet.iter_rows()):
    if row[7] != None:
        if row[7].hyperlink:
            string_name = str(row[7].value)
            hyperlinks[string_name] = row[7].hyperlink.target


# Iterate through rows and append to data
for row in worksheet.iter_rows():
    row_data = []
    for cell in row:
        if cell.value is None:
            row_data.append("")
        else:
            row_data.append(cell.value)
    # check if the row_data[0] contains a week number if not use previous
    if not row_data[0]:
        row_data[0] = week_num
    else:
        week_num = row_data[0]
    data.append(row_data)

# replace links with their url
for i,row in enumerate(data):
    if i != 0:
        if row[7] != None and row[7] != "" and row[7] != str():
            data[i][7] = hyperlinks[str(row[7])]

#Reformat the data
output = {"weeks": [{"week_num":week[0],
                    week[1].lower():[{"Content":week[2]},
                                       {"pages":week[3]},
                                       {"task":week[4]},
                                       {"task_lower":week[5]},
                                       {"task_higher":week[6]},
                                       {"links":week[7].split(" ")},
                                       {"info":week[8]}]
                   } for week in data[1:] ]}